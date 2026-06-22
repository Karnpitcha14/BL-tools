from flask import Flask, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import io
import os
import sys

app = Flask(__name__)

def get_html():
    if hasattr(sys, '_MEIPASS'):
        path = os.path.join(sys._MEIPASS, 'index.html')
    else:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html')
    return open(path, encoding='utf-8').read()

HTML = get_html()

@app.route('/')
def index():
    return HTML

def parse_line_text(text):
    """Parse LINE message to extract SKU, qty, team using regex — no API needed"""
    items = []
    current_team = 'UNKNOWN'

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        # Detect team header
        line_upper = line.upper()
        if re.search(r'ท[ีi]ม\s*FO|^FO\s*:|^FO$|FO\s+สรุป|สรุป.*FO', line_upper):
            current_team = 'FO'
            continue
        if re.search(r'ท[ีi]ม\s*MC|^MC\s*:|^MC$|MC\s+สรุป|สรุป.*MC', line_upper):
            current_team = 'MC'
            continue

        # Match SKU patterns:
        # 030100296=1, 030100296=1 ชิ้น
        # 030100296 (1 ชิ้น), 030100296 (1)
        # 030100296 จำนวน 1 ชิ้น, 030100296 จำนวน 1
        # 030100296 1 ชิ้น
        sku_pattern = r'(\d[\d\s]{5,14}\d)'
        qty_pattern = r'[=\s\(]+(\d+)'

        sku_match = re.search(sku_pattern, line)
        if not sku_match:
            continue

        sku = re.sub(r'\s+', '', sku_match.group(1))  # remove spaces in SKU

        # Find qty after SKU
        rest = line[sku_match.end():]
        qty_match = re.search(qty_pattern, rest)
        if not qty_match:
            continue

        qty = int(qty_match.group(1))

        # Detect inline team override
        team = current_team
        if re.search(r'\bFO\b', line_upper):
            team = 'FO'
        elif re.search(r'\bMC\b', line_upper):
            team = 'MC'

        items.append({'sku': sku, 'qty': qty, 'team': team})

    return items

@app.route('/api/parse-line', methods=['POST'])
def parse_line():
    text = request.json.get('text', '')
    try:
        items = parse_line_text(text)
        return jsonify({"ok": True, "items": items})
    except Exception as e:
        return jsonify({"ok": False, "items": [], "error": str(e)})

@app.route('/api/check-stock', methods=['POST'])
def check_stock():
    stock_file = request.files.get('stock')
    order_file = request.files.get('order')
    items_json = request.form.get('items', '[]')
    items = json.loads(items_json)

    if not stock_file or not items:
        return jsonify({"ok": False, "error": "ไม่มีไฟล์หรือ SKU"})

    df_stock = pd.read_excel(stock_file, header=1, dtype={'Seller SKU': str})

    # อ่านไฟล์ออเดอร์แบบประหยัด RAM — อ่านแค่คอลัมน์ที่จำเป็น
    df_order = None
    if order_file:
        # อ่านครั้งแรกเพื่อดูชื่อคอลัมน์
        cols_preview = pd.read_excel(order_file, dtype=str, nrows=0)
        all_cols = list(cols_preview.columns)
        order_file.seek(0)
        df_order = pd.read_excel(order_file, dtype=str, engine='openpyxl')

    # หา SKU col ในไฟล์ออเดอร์
    order_sku_col = None
    if df_order is not None:
        for col in ['Fscode', 'รหัสสินค้า', 'SKU', 'sku']:
            if col in df_order.columns:
                order_sku_col = col
                break

    all_yokro_skus = {it['sku'] for it in items}

    wb = openpyxl.Workbook()
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    hfill = PatternFill("solid", start_color="E91E63", end_color="E91E63")
    gfill = PatternFill("solid", start_color="C8E6C9", end_color="C8E6C9")
    yfill = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")
    ofill = PatternFill("solid", start_color="FFE0B2", end_color="FFE0B2")
    rfill = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")

    def hdr(ws, r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def dc(ws, r, c, v, fill, left=False, red=False, bold=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=10, color="FF0000" if red else "000000", bold=bold)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left" if left else "center", vertical="center", wrap_text=True)
        cell.border = border

    # Sheet 1: เช็คสต็อก
    ws1 = wb.active
    ws1.title = "เช็คสต็อกยกรอ"
    ws1.freeze_panes = 'A2'
    headers1 = ["ทีม","SKU","ชื่อสินค้า","MC","SH","Total","แจ้งยกรอ","รอแพ็คจริง","ผล","หมายเหตุ"]
    for ci,h in enumerate(headers1,1): hdr(ws1,1,ci,h)

    for ri, it in enumerate(items, 2):
        sku, rep, team = it['sku'], it['qty'], it['team']
        row = df_stock[df_stock['Seller SKU'] == sku]
        if row.empty:
            vals = [team,sku,"ไม่พบ SKU","-","-","-",rep,0,"❓ ไม่พบ","ไม่พบ SKU นี้ในระบบ"]
            fill = rfill
        else:
            r = row.iloc[0]
            mc,sh,total = int(r['MC']),int(r['SH']),int(r['Total'])
            ship = int(r['Shipment Unit']) if pd.notna(r['Shipment Unit']) else 0
            wait = int(r['Allowcate Unit']) - ship
            desc = str(r['Description'])[:45]
            if wait == 0:
                result,fill,note = "❌ ไม่มีรอแพ็ค",rfill,"มีของในคลัง แต่ไม่มี Allocate ค้าง"
            elif wait == rep:
                result,fill,note = "✅ ตรง",gfill,"ตรงกับที่แจ้ง"
            elif wait > rep:
                result,fill,note = "⚠️ รอแพ็คมากกว่า",yfill,f"รอแพ็คจริง {wait} > แจ้ง {rep}"
            else:
                result,fill,note = "⚠️ รอแพ็คน้อยกว่า",ofill,f"รอแพ็คจริง {wait} < แจ้ง {rep}"
            vals = [team,sku,desc,mc or'-',sh or'-',total,rep,wait,result,note]
        for ci,val in enumerate(vals,1):
            dc(ws1,ri,ci,val,fill,left=(ci==3))

    for ci,w in enumerate([8,14,38,8,8,8,12,12,18,28],1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2: Detail ออเดอร์
    if df_order is not None and order_sku_col:
        ws2 = wb.create_sheet("Detail ออเดอร์")
        ws2.freeze_panes = 'A2'
        pfill = PatternFill("solid", start_color="FCE4EC", end_color="FCE4EC")
        wfill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        ord_cols = list(df_order.columns) + ['หมายเหตุ']
        for ci,h in enumerate(ord_cols,1): hdr(ws2,1,ci,h)

        # ทำ team map
        team_map = {it['sku']: it['team'] for it in items}

        for ri,(_, row) in enumerate(df_order.iterrows(),2):
            sku = str(row.get(order_sku_col,'')).strip()
            is_yokro = sku in all_yokro_skus
            team = team_map.get(sku,'')
            remark = f"ยกรอ ({team})" if is_yokro else ''
            base_fill = wfill if ri%2==0 else pfill
            for ci,col in enumerate(ord_cols,1):
                val = remark if col=='หมายเหตุ' else (row[col] if col in row.index and pd.notna(row[col]) else '')
                dc(ws2,ri,ci,val,base_fill,red=is_yokro,bold=(col=='หมายเหตุ' and is_yokro),
                   left=(col in ['Description','รายละเอียด','Campaign']))
        ws2.column_dimensions['A'].width = 16
        for ci in range(2, len(ord_cols)+1):
            ws2.column_dimensions[get_column_letter(ci)].width = 16
        if 'Description' in ord_cols:
            ci = ord_cols.index('Description')+1
            ws2.column_dimensions[get_column_letter(ci)].width = 38
        if 'รายละเอียด' in ord_cols:
            ci = ord_cols.index('รายละเอียด')+1
            ws2.column_dimensions[get_column_letter(ci)].width = 38

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="สรุปยกรอ.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/oos', methods=['POST'])
def oos():
    try:
        oos_file   = request.files.get('oos')
        stock_file = request.files.get('stock')
        recv_file  = request.files.get('receive')
        if not oos_file or not stock_file or not recv_file:
            return jsonify({"ok": False, "error": "ต้องอัปโหลดครบ 3 ไฟล์"})

        df_oos   = pd.read_excel(oos_file, dtype=str)
        df_stock = pd.read_excel(stock_file, header=1, dtype={'Seller SKU': str})
        df_recv  = pd.read_excel(recv_file, dtype=str)

        # --- กรอง OOS: TikTok + Shopee เท่านั้น + ไม่ใช่ Preorder (ลงท้าย P) ---
        trans_col = next((c for c in df_oos.columns if 'trans' in c.lower() or 'source' in c.lower()), None)
        camp_col  = next((c for c in df_oos.columns if 'campaign' in c.lower()), None)
        sku_col   = next((c for c in df_oos.columns if c.lower() in ['fscode','sku','seller sku','รหัสสินค้า']), None)
        name_col  = next((c for c in df_oos.columns if 'description' in c.lower() or 'รายละเอียด' in c.lower() or 'ชื่อ' in c.lower()), None)

        if trans_col:
            df_oos = df_oos[df_oos[trans_col].str.lower().isin(['tiktok_shop','shopee'])]
        if camp_col:
            df_oos = df_oos[~df_oos[camp_col].astype(str).str.upper().str.endswith('P')]

        if sku_col is None:
            return jsonify({"ok": False, "error": "ไม่พบคอลัมน์ SKU ในไฟล์ OOS"})

        # --- นับ OOS ค้างต่อ SKU ---
        qty_col = next((c for c in df_oos.columns if 'unit' in c.lower() or 'จำนวน' in c.lower() or 'qty' in c.lower()), None)
        if qty_col:
            oos_summary = df_oos.groupby(sku_col).agg(
                oos_qty=(qty_col, lambda x: pd.to_numeric(x, errors='coerce').sum()),
                name=(name_col if name_col else sku_col, 'first')
            ).reset_index()
        else:
            oos_summary = df_oos.groupby(sku_col).agg(
                oos_qty=(sku_col, 'count'),
                name=(name_col if name_col else sku_col, 'first')
            ).reset_index()

        # --- คำนวณคงเหลือจริง = Total - Allowcate Unit (vectorized) ---
        df_stock['_sku']    = df_stock['Seller SKU'].astype(str)
        df_stock['_total']  = pd.to_numeric(df_stock['Total'], errors='coerce').fillna(0).astype(int)
        df_stock['_alloc']  = pd.to_numeric(df_stock['Allowcate Unit'], errors='coerce').fillna(0).astype(int)
        df_stock['_remain'] = df_stock['_total'] - df_stock['_alloc']
        df_stock['_desc']   = df_stock['Description'].astype(str).str[:40]
        stock_df = df_stock[['_sku','_remain','_desc']].drop_duplicates('_sku').set_index('_sku')
        stock_map = {sku: {'remain': int(row['_remain']), 'desc': row['_desc']}
                     for sku, row in stock_df.iterrows()}

        # --- ใบรับเข้า (vectorized) ---
        recv_sku_col    = next((c for c in df_recv.columns if 'sku' in c.lower() or 'รหัส' in c.lower()), None)
        recv_status_col = next((c for c in df_recv.columns if 'สถานะ' in c.lower() or 'status' in c.lower()), None)
        recv_doc_col    = next((c for c in df_recv.columns if 'เลข' in c.lower() or 'เอกสาร' in c.lower() or 'doc' in c.lower()), None)

        recv_map = {}
        if recv_sku_col:
            df_recv['_rsku']   = df_recv[recv_sku_col].astype(str).str.strip()
            df_recv['_status'] = df_recv[recv_status_col].astype(str).str.strip() if recv_status_col else ''
            df_recv['_doc']    = df_recv[recv_doc_col].astype(str).str.strip() if recv_doc_col else ''
            for row in df_recv[['_rsku','_status','_doc']].to_dict('records'):
                sku = row['_rsku']
                if sku not in recv_map:
                    recv_map[sku] = []
                recv_map[sku].append({'status': row['_status'], 'doc_no': row['_doc']})

        # --- จัดกลุ่ม ---
        group_a, group_b, group_c, group_e = [], [], [], []

        for _, row in oos_summary.iterrows():
            sku      = str(row[sku_col])
            oos_qty  = int(row['oos_qty'])
            name     = str(row['name'])[:40]
            remain   = stock_map.get(sku, {}).get('remain', 0)
            if not name or name == sku:
                name = stock_map.get(sku, {}).get('desc', sku)

            # F: ของพอ
            if remain >= oos_qty:
                continue

            recv_entries = recv_map.get(sku, [])
            has_unconfirm = any('ยังไม่' in e['status'] or 'unconfirm' in e['status'].lower() for e in recv_entries)
            has_confirm   = any('ยืนยันแล้ว' in e['status'] or 'confirm' in e['status'].lower() for e in recv_entries)
            doc_no_a      = next((e['doc_no'] for e in recv_entries if 'ยังไม่' in e['status'] or 'unconfirm' in e['status'].lower()), '')

            if has_unconfirm:
                group_a.append({'sku': sku, 'name': name, 'doc_no': doc_no_a})
            elif has_confirm:
                group_b.append({'sku': sku, 'name': name})
            elif remain <= 0:
                group_c.append({'sku': sku, 'name': name})
            else:
                group_e.append({'sku': sku, 'name': name, 'remain': remain, 'oos_qty': oos_qty})

        # --- สร้างข้อความ ---
        def fmt_list(items, with_num=True):
            return '\n'.join(f"{i+1}. {it['sku']} {it['name']}" if with_num else f"{it['sku']} {it['name']}"
                             for i, it in enumerate(items))

        block_a, block_b, block_c, block_e = None, None, None, None

        if group_a:
            # Group by doc_no
            from collections import defaultdict
            by_doc = defaultdict(list)
            for it in group_a:
                by_doc[it['doc_no']].append(it)
            lines = ['แจ้งรายการรับและเบิกด่วนนะคะ @เค @อาร์ม']
            for doc, items in by_doc.items():
                if doc:
                    lines.append(f'\n[{doc}]')
                for i, it in enumerate(items):
                    lines.append(f"{i+1}. {it['sku']} {it['name']}")
            lines.append('\nขอบคุณค่ะ')
            block_a = '\n'.join(lines)

        if group_b:
            lines = ['รบกวนเบิกรายการนี้ด่วนนะคะ @เค @อาร์ม (รับเข้าแล้ว)']
            for i, it in enumerate(group_b):
                lines.append(f"{i+1}. {it['sku']} {it['name']}")
            lines.append('\nขอบคุณค่ะ')
            block_b = '\n'.join(lines)

        if group_c:
            lines = [f"{it['sku']} {it['name']}" for it in group_c]
            lines.append('\nรายการที่ไม่มีสินค้า และยังไม่ได้เปิดใบรับเข้านะคะพี่ป๊อก')
            block_c = '\n'.join(lines)

        if group_e:
            lines = [f"{it['sku']} {it['name']}  (คงเหลือจริง {it['remain']} / OOS ค้าง {it['oos_qty']})"
                     for it in group_e]
            lines.append('\nรายการที่สินค้าไม่เพียงพอ และยังไม่ได้เปิดใบรับเข้านะคะพี่ป๊อก')
            block_e = '\n'.join(lines)

        return jsonify({"ok": True, "blockA": block_a, "blockB": block_b, "blockC": block_c, "blockE": block_e})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


if __name__ == '__main__':
    app.run(port=5050, debug=False, use_reloader=False)
